"""
Report Generator — รายการคำนวณโครงสร้างเหล็ก
สร้างเอกสาร PDF มาตรฐานสำหรับยื่นขออนุญาตก่อสร้าง

ฟอร์แมต: A4, ภาษาไทย/อังกฤษ, ตาม วสท. 011038-22
รองรับ: คาน (Beam), เสา (Column), แป (Purlin),
        รอยต่อ (Connection), แผ่นฐาน (Base Plate)
"""
from __future__ import annotations

import io
import math
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Union

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate, Frame, HRFlowable, PageTemplate,
    Paragraph, Spacer, Table, TableStyle, KeepTogether,
)
from reportlab.platypus.flowables import Flowable

# Matplotlib for diagram generation (optional)
try:
    import matplotlib
    matplotlib.use('Agg')  # Non-GUI backend
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Excel export (optional)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    EXCELPY_AVAILABLE = True
except ImportError:
    EXCELPY_AVAILABLE = False

# ── Font registration ────────────────────────────────────────────────────────
_FONT_DIR = "C:/Windows/Fonts"
_FONTS_REGISTERED = False


def _register_fonts():
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    candidates = [
        ("ARIALUNI.ttf",  "Thai"),
        ("arialbd.ttf",   "ThaiBd"),
        ("arial.ttf",     "ThaiReg"),
        ("tahoma.ttf",    "ThaiFallback"),
    ]
    for fname, name in candidates:
        path = os.path.join(_FONT_DIR, fname)
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
            except Exception:
                pass
    _FONTS_REGISTERED = True


# ── Color palette ─────────────────────────────────────────────────────────────
C_NAVY   = colors.HexColor("#0d1b2a")
C_BLUE   = colors.HexColor("#1b4f72")
C_ACCENT = colors.HexColor("#2e86ab")
C_LIGHT  = colors.HexColor("#f0f4f8")
C_PASS   = colors.HexColor("#1a7a4a")
C_FAIL   = colors.HexColor("#c0392b")
C_WARN   = colors.HexColor("#d68910")
C_PASS_BG = colors.HexColor("#d4f5e4")
C_FAIL_BG = colors.HexColor("#fde8e8")
C_WARN_BG = colors.HexColor("#fff3cd")
C_ROW_ALT = colors.HexColor("#f7fafc")
C_BORDER  = colors.HexColor("#dde3ec")
C_WHITE   = colors.white
C_BLACK   = colors.black


# ── Project info ─────────────────────────────────────────────────────────────
@dataclass
class ProjectInfo:
    """Project metadata shown on cover and every page header."""
    project_name: str = "โครงการ"
    project_no:   str = ""
    engineer:     str = ""
    checker:      str = ""
    date:         str = ""          # auto-filled if empty
    standard:     str = "วสท. 011038-22"
    method:       str = "ASD (Allowable Stress Design)"
    client:       str = ""
    location:     str = ""

    def __post_init__(self):
        if not self.date:
            self.date = datetime.now().strftime("%d/%m/%Y")


# ── Style factory ─────────────────────────────────────────────────────────────

def _styles() -> Dict[str, ParagraphStyle]:
    _register_fonts()
    base = "Thai"
    bold = "ThaiBd"

    def S(name, **kw) -> ParagraphStyle:
        return ParagraphStyle(name, **kw)

    return {
        "cover_title": S("cover_title", fontName=bold,   fontSize=16,
                          textColor=C_WHITE, alignment=TA_CENTER, leading=22),
        "cover_sub":   S("cover_sub",   fontName=base,   fontSize=10,
                          textColor=colors.HexColor("#a8d8ea"), alignment=TA_CENTER, leading=14),
        "h1":          S("h1",          fontName=bold,   fontSize=12,
                          textColor=C_WHITE,  spaceAfter=0, spaceBefore=4, leading=16),
        "h2":          S("h2",          fontName=bold,   fontSize=10,
                          textColor=C_BLUE,  spaceAfter=3, spaceBefore=8, leading=14),
        "h3":          S("h3",          fontName=bold,   fontSize=9,
                          textColor=C_NAVY,  spaceAfter=2, spaceBefore=4, leading=12),
        "body":        S("body",        fontName=base,   fontSize=9,
                          textColor=C_NAVY,  spaceAfter=2, leading=13),
        "body_bold":   S("body_bold",   fontName=bold,   fontSize=9,
                          textColor=C_NAVY,  leading=13),
        "mono":        S("mono",        fontName="Courier", fontSize=8.5,
                          textColor=C_NAVY,  leading=12),
        "cell":        S("cell",        fontName=base,   fontSize=8.5,
                          textColor=C_NAVY,  leading=11),
        "cell_bold":   S("cell_bold",   fontName=bold,   fontSize=8.5,
                          textColor=C_NAVY,  leading=11),
        "cell_hdr":    S("cell_hdr",    fontName=bold,   fontSize=8.5,
                          textColor=C_WHITE, alignment=TA_CENTER, leading=11),
        "cell_r":      S("cell_r",      fontName=base,   fontSize=8.5,
                          textColor=C_NAVY,  alignment=TA_RIGHT, leading=11),
        "cell_c":      S("cell_c",      fontName=base,   fontSize=8.5,
                          textColor=C_NAVY,  alignment=TA_CENTER, leading=11),
        "pass":        S("pass",        fontName=bold,   fontSize=9,
                          textColor=C_PASS,  leading=12),
        "fail":        S("fail",        fontName=bold,   fontSize=9,
                          textColor=C_FAIL,  leading=12),
        "warn":        S("warn",        fontName=bold,   fontSize=9,
                          textColor=C_WARN,  leading=12),
        "footnote":    S("footnote",    fontName=base,   fontSize=7.5,
                          textColor=colors.grey, leading=10),
        "formula_lbl": S("formula_lbl", fontName=bold,    fontSize=8.5,
                          textColor=C_BLUE,  leading=11),
        "formula_eq":  S("formula_eq",  fontName="Courier", fontSize=8.5,
                          textColor=C_NAVY,  leading=11),
        "formula_val": S("formula_val", fontName="Courier-Bold", fontSize=8.5,
                          textColor=C_NAVY,  alignment=TA_RIGHT, leading=11),
        "result_big":  S("result_big",  fontName=bold,   fontSize=11,
                          textColor=C_NAVY,  alignment=TA_CENTER, leading=15),
    }


# ── Helper flowables ──────────────────────────────────────────────────────────

class _SectionHeader(Flowable):
    """Colored banner for section headings (e.g. '1. คุณสมบัติหน้าตัด')."""
    def __init__(self, text: str, width: float = 16.7 * cm,
                 bg=C_BLUE, fg=C_WHITE):
        super().__init__()
        self._text = text
        self._width = width
        self._bg = bg
        self._fg = fg
        self.height = 18

    def draw(self):
        c = self.canv
        c.setFillColor(self._bg)
        c.roundRect(0, 0, self._width, 16, 3, fill=1, stroke=0)
        c.setFillColor(self._fg)
        c.setFont("ThaiBd", 9.5)
        c.drawString(6, 4, self._text)

    def wrap(self, *_):
        return self._width, 18


class _SubHeader(Flowable):
    """Light accent strip for sub-sections."""
    def __init__(self, text: str, width: float = 16.7 * cm):
        super().__init__()
        self._text = text
        self._width = width
        self.height = 15

    def draw(self):
        c = self.canv
        c.setFillColor(C_LIGHT)
        c.rect(0, 0, self._width, 13, fill=1, stroke=0)
        c.setStrokeColor(C_ACCENT)
        c.setLineWidth(2.5)
        c.line(0, 0, 0, 13)
        c.setFillColor(C_BLUE)
        c.setFont("ThaiBd", 8.5)
        c.drawString(7, 3, self._text)

    def wrap(self, *_):
        return self._width, 15


class _PassFailBanner(Flowable):
    """Large PASS/FAIL result banner."""
    def __init__(self, is_ok: bool, status_th: str, width: float = 16.7 * cm):
        super().__init__()
        self._ok = is_ok
        self._status = status_th
        self._width = width
        self.height = 32

    def draw(self):
        c = self.canv
        bg = C_PASS_BG if self._ok else C_FAIL_BG
        fg = C_PASS    if self._ok else C_FAIL
        label = "✓  ผ่าน (ADEQUATE)" if self._ok else "✗  ไม่ผ่าน (INADEQUATE)"
        c.setFillColor(bg)
        c.roundRect(0, 0, self._width, 30, 4, fill=1, stroke=0)
        c.setStrokeColor(fg)
        c.setLineWidth(1.5)
        c.roundRect(0, 0, self._width, 30, 4, fill=0, stroke=1)
        c.setFillColor(fg)
        c.setFont("ThaiBd", 12)
        c.drawString(10, 16, label)
        c.setFont("Thai", 8)
        c.drawString(10, 5, self._status)

    def wrap(self, *_):
        return self._width, 32


# ── Table helpers ─────────────────────────────────────────────────────────────

def _ts_base(col_widths) -> TableStyle:
    """Base table style with borders and alternating rows."""
    return TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  C_BLUE),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  C_WHITE),
        ("FONTNAME",    (0, 0), (-1, 0),  "ThaiBd"),
        ("FONTSIZE",    (0, 0), (-1, 0),  8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_ROW_ALT]),
        ("FONTNAME",    (0, 1), (-1, -1), "Thai"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8.5),
        ("GRID",        (0, 0), (-1, -1), 0.4, C_BORDER),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0),(-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",(0, 0), (-1, -1), 5),
    ])


def _t(data, col_widths, extra_style=None, hdr_align="CENTER") -> Table:
    """Build a styled table from data rows."""
    st = _ts_base(col_widths)
    if hdr_align == "LEFT":
        st.add("ALIGN", (0, 0), (-1, 0), "LEFT")
    else:
        st.add("ALIGN", (0, 0), (-1, 0), "CENTER")
    if extra_style:
        for cmd in extra_style:
            st.add(*cmd)
    tbl = Table(data, colWidths=col_widths, style=st, repeatRows=1)
    return tbl


def _ratio_style(ratio: float) -> Tuple[colors.Color, colors.Color]:
    """Return (bg, fg) for a ratio value cell."""
    if ratio <= 0.85:
        return C_PASS_BG, C_PASS
    elif ratio <= 1.0:
        return C_WARN_BG, C_WARN
    else:
        return C_FAIL_BG, C_FAIL


def _fmt(v, d=2) -> str:
    if v is None:
        return "—"
    if abs(v) == float("inf") or math.isnan(v):
        return "—"
    return f"{v:,.{d}f}"


# ── Page template ─────────────────────────────────────────────────────────────

class _ReportDoc(BaseDocTemplate):
    """A4 document with running header/footer on every page."""

    def __init__(self, buf, project: ProjectInfo):
        self._proj = project
        self._page_count = 0
        super().__init__(
            buf,
            pagesize=A4,
            leftMargin=20 * mm,
            rightMargin=20 * mm,
            topMargin=28 * mm,
            bottomMargin=22 * mm,
            title=f"รายการคำนวณ — {project.project_name}",
            author=project.engineer or "โปรแกรมออกแบบโครงสร้างเหล็ก",
        )
        frame = Frame(
            self.leftMargin, self.bottomMargin,
            self.width, self.height,
            id="main", showBoundary=0,
        )
        self.addPageTemplates([PageTemplate(id="main", frames=[frame],
                                            onPage=self._draw_page)])

    def _draw_page(self, canvas, doc):
        canvas.saveState()
        W, H = A4
        # ── Top bar ──────────────────────────────────────────────────
        canvas.setFillColor(C_NAVY)
        canvas.rect(0, H - 18 * mm, W, 18 * mm, fill=1, stroke=0)
        canvas.setFillColor(C_WHITE)
        canvas.setFont("ThaiBd", 9)
        canvas.drawString(20 * mm, H - 10 * mm,
                          f"รายการคำนวณออกแบบโครงสร้างเหล็ก  |  {self._proj.project_name}")
        canvas.setFont("Thai", 8)
        canvas.setFillColor(colors.HexColor("#a8d8ea"))
        canvas.drawRightString(W - 20 * mm, H - 10 * mm,
                               f"{self._proj.standard}  |  {self._proj.method}")
        # ── Bottom bar ───────────────────────────────────────────────
        canvas.setFillColor(C_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=1, stroke=0)
        canvas.setStrokeColor(C_BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(20 * mm, 14 * mm, W - 20 * mm, 14 * mm)
        canvas.setFillColor(C_BLUE)
        canvas.setFont("Thai", 7.5)
        canvas.drawString(20 * mm, 5 * mm,
                          f"ออกแบบโดย: {self._proj.engineer or '—'}   "
                          f"ตรวจสอบโดย: {self._proj.checker or '—'}   "
                          f"วันที่: {self._proj.date}")
        canvas.setFont("ThaiBd", 7.5)
        canvas.drawRightString(W - 20 * mm, 5 * mm,
                               f"หน้า {doc.page}")
        canvas.restoreState()


# ── Cover page ────────────────────────────────────────────────────────────────

def _cover_elements(project: ProjectInfo, module_title: str,
                    stories: List[str]) -> List:
    """Build cover page flowables."""
    ST = _styles()
    W = 16.7 * cm
    elems: List = []

    elems.append(Spacer(1, 3 * cm))

    # Title box
    title_data = [
        [Paragraph("รายการคำนวณ", ST["cover_title"])],
        [Paragraph(module_title, ST["cover_title"])],
        [Paragraph(f"โปรแกรมออกแบบโครงสร้างเหล็ก (วสท. 011038-22)", ST["cover_sub"])],
    ]
    title_tbl = Table(title_data, colWidths=[W])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C_NAVY),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
    ]))
    elems.append(title_tbl)
    elems.append(Spacer(1, 1.5 * cm))

    # Project info table
    info_rows = [
        ["ชื่อโครงการ", project.project_name],
        ["เลขที่โครงการ", project.project_no or "—"],
        ["ผู้รับผิดชอบ", project.client or "—"],
        ["สถานที่", project.location or "—"],
        ["ออกแบบโดย", project.engineer or "—"],
        ["ตรวจสอบโดย", project.checker or "—"],
        ["วันที่", project.date],
        ["มาตรฐาน", project.standard],
        ["วิธีการออกแบบ", project.method],
    ]
    tbl_data = [[Paragraph(k, ST["cell_bold"]), Paragraph(v, ST["cell"])]
                for k, v in info_rows]
    info_tbl = Table(tbl_data, colWidths=[4.5 * cm, 12.2 * cm])
    info_tbl.setStyle(TableStyle([
        ("GRID",       (0, 0), (-1, -1), 0.4, C_BORDER),
        ("BACKGROUND", (0, 0), (0, -1),  C_LIGHT),
        ("FONTNAME",   (0, 0), (-1, -1), "Thai"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 7),
    ]))
    elems.append(info_tbl)

    if stories:
        elems.append(Spacer(1, 1 * cm))
        elems.append(_SubHeader("รายการคำนวณที่รวมอยู่ในเอกสาร", W))
        for i, s in enumerate(stories, 1):
            elems.append(Paragraph(f"  {i}.  {s}", ST["body"]))

    return elems


# ── Formula table builder ─────────────────────────────────────────────────────

def _formula_table(rows: List[Tuple[str, str, str, str]],
                   col_w: List[float] = None) -> Table:
    """
    Build a formula/calculation table.
    rows: list of (label, formula, substitution, result_with_unit)
    col_w: column widths in cm
    """
    ST = _styles()
    if col_w is None:
        col_w = [4.5 * cm, 4.5 * cm, 5.5 * cm, 2.2 * cm]
    header = [
        Paragraph("ตัวแปร", ST["cell_hdr"]),
        Paragraph("สูตร", ST["cell_hdr"]),
        Paragraph("แทนค่า", ST["cell_hdr"]),
        Paragraph("ผลลัพธ์", ST["cell_hdr"]),
    ]
    data = [header]
    for i, (lbl, formula, sub, result) in enumerate(rows):
        bg = C_WHITE if i % 2 == 0 else C_ROW_ALT
        data.append([
            Paragraph(lbl,     ST["formula_lbl"]),
            Paragraph(formula, ST["formula_eq"]),
            Paragraph(sub,     ST["formula_eq"]),
            Paragraph(result,  ST["formula_val"]),
        ])
    tbl = Table(data, colWidths=col_w)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  C_BLUE),
        ("GRID",         (0, 0), (-1, -1), 0.4, C_BORDER),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_ROW_ALT]),
    ]))
    return tbl


def _check_row_style(tbl, row_idx: int, ratio: float):
    """Add pass/fail background to a table row."""
    bg, _ = _ratio_style(ratio)
    tbl._argW  # access to force build
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, row_idx), (-1, row_idx), bg),
    ]))


# ── Section properties table ──────────────────────────────────────────────────

def _section_props_table(props: Dict[str, Any]) -> Table:
    """Build a 2-column section properties table from a dict."""
    ST = _styles()
    items = list(props.items())
    # 2 items per row → 4-column table
    rows = []
    for i in range(0, len(items), 2):
        k1, v1 = items[i]
        if i + 1 < len(items):
            k2, v2 = items[i + 1]
        else:
            k2, v2 = "", ""
        rows.append([
            Paragraph(k1, ST["cell_bold"]),
            Paragraph(str(v1), ST["cell_r"]),
            Paragraph(k2, ST["cell_bold"]),
            Paragraph(str(v2), ST["cell_r"]),
        ])
    tbl = Table(rows, colWidths=[3.8 * cm, 2.5 * cm, 3.8 * cm, 2.5 * cm])
    tbl.setStyle(TableStyle([
        ("GRID",       (0, 0), (-1, -1), 0.4, C_BORDER),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [C_WHITE, C_ROW_ALT]),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, 0), (0, -1), C_LIGHT),
        ("BACKGROUND", (2, 0), (2, -1), C_LIGHT),
    ]))
    return tbl


# ── Load combination check table ──────────────────────────────────────────────

def _load_combo_table(load_cases: List[Dict], cols: List[str],
                      keys: List[str], units: List[str],
                      ratio_key: str = "stress_ratio") -> Table:
    ST = _styles()
    W_TOTAL = 16.7 * cm
    n_cols = len(cols)
    col_w = [W_TOTAL / n_cols] * n_cols

    header = [Paragraph(c, ST["cell_hdr"]) for c in cols]
    data = [header]
    styles_extra = []

    for row_i, lc in enumerate(load_cases, 1):
        ratio = lc.get(ratio_key, 0)
        bg, _ = _ratio_style(ratio) if ratio else (C_WHITE, C_NAVY)
        row = []
        for k in keys:
            v = lc.get(k, "")
            if isinstance(v, float):
                row.append(Paragraph(_fmt(v, 3), ST["cell_r"]))
            else:
                row.append(Paragraph(str(v), ST["cell"]))
        data.append(row)
        styles_extra.append(("BACKGROUND", (0, row_i), (-1, row_i), bg))

    tbl = Table(data, colWidths=col_w)
    base = _ts_base(col_w)
    for cmd in styles_extra:
        base.add(*cmd)
    tbl.setStyle(base)
    return tbl


# ══════════════════════════════════════════════════════════════════════════════
# BEAM REPORT
# ══════════════════════════════════════════════════════════════════════════════

def _beam_elements(result, section_name: str, span: float) -> List:
    """Build beam calculation flowables."""
    from beam_design import BeamDesignResult
    ST = _styles()
    W = 16.7 * cm
    r: BeamDesignResult = result
    props = r.details.get("properties", {}) if r.details else {}
    Fy = props.get("Fy", 0)
    Sx = props.get("Sx", 0)
    E = 2.0e5
    L_mm = span * 1000

    elems: List = []
    elems.append(_SectionHeader("1.  คุณสมบัติหน้าตัด (Section Properties)", W))
    elems.append(Spacer(1, 3))

    sec_props = {
        "หน้าตัด (Section)":     section_name,
        "ช่วงคาน L":              f"{_fmt(span)} m",
        "กำลังคราก Fy":           f"{_fmt(Fy, 0)} MPa",
        "โมดูลัสหน้าตัด Sx":     f"{_fmt(Sx / 1e3, 0)} × 10³ mm³",
        "โมเมนต์อินีเชีย Ix":    f"{_fmt(props.get('Ix', 0) / 1e6, 2) if props.get('Ix') else '—'} × 10⁶ mm⁴",
        "E (เหล็ก)":              "200,000 MPa",
        "หน่วยแรงดัดที่ยอมให้ Fb": f"{_fmt(r.Fb)} MPa",
        "หน่วยแรงเฉือนที่ยอมให้ Fv": f"{_fmt(r.Fv)} MPa",
    }
    elems.append(_section_props_table(sec_props))
    elems.append(Spacer(1, 5))

    # ── 2. Load combinations ──────────────────────────────────────────
    elems.append(_SectionHeader("2.  การรวมน้ำหนักบรรทุก (Load Combinations — ASD)", W))
    elems.append(Spacer(1, 3))
    if r.details and r.details.get("load_cases"):
        lc_cols = ["กรณีน้ำหนัก", "w (kN/m)", "M (kN-m)", "V (kN)",
                   "fb (MPa)", "fv (MPa)", "fb/Fb", "fv/Fv"]
        lc_keys = ["name", "w_kN_m", "M_kNm", "V_kN", "fb_MPa", "fv_MPa",
                   "stress_ratio", "shear_ratio"]
        lc_units = ["", "kN/m", "kN-m", "kN", "MPa", "MPa", "", ""]
        elems.append(_load_combo_table(r.details["load_cases"], lc_cols,
                                       lc_keys, lc_units, "stress_ratio"))
    elems.append(Spacer(1, 5))

    # ── 3. Bending check formulas ─────────────────────────────────────
    elems.append(_SectionHeader("3.  การตรวจสอบหน่วยแรงดัด (Bending Stress Check)", W))
    elems.append(Spacer(1, 3))
    elems.append(_SubHeader(f"กรณีวิกฤต: {r.critical_load_case}", W))
    elems.append(Spacer(1, 2))

    w_crit = r.max_moment * 8 / (span ** 2) if span > 0 else 0   # kN/m back-calc
    bending_rows = [
        ("M_max",  "M = w·L²/8",
         f"= {_fmt(w_crit)}×{_fmt(span)}²/8",
         f"{_fmt(r.max_moment)} kN-m"),
        ("fb = M/Sx",
         "fb = M×10⁶/Sx",
         f"= {_fmt(r.max_moment*1e6/Sx if Sx else 0)}",
         f"{_fmt(r.fb)} MPa"),
        ("Fb (ยอมให้)",
         "Fb = 0.66·Fy",
         f"= 0.66×{_fmt(Fy, 0)}",
         f"{_fmt(r.Fb)} MPa"),
        ("fb/Fb ≤ 1.0",
         "อัตราส่วนการใช้งาน",
         f"= {_fmt(r.fb)}/{_fmt(r.Fb)}",
         f"{_fmt(r.stress_ratio, 3)}"),
    ]
    ft = _formula_table(bending_rows)
    elems.append(ft)
    elems.append(Spacer(1, 5))

    # ── 4. Shear check ────────────────────────────────────────────────
    elems.append(_SectionHeader("4.  การตรวจสอบหน่วยแรงเฉือน (Shear Stress Check)", W))
    elems.append(Spacer(1, 3))
    shear_rows = [
        ("V_max",  "V = w·L/2",
         f"= {_fmt(w_crit)}×{_fmt(span)}/2",
         f"{_fmt(r.max_shear)} kN"),
        ("fv = V/(d·tw)",
         "fv = V×10³/(d·tw)",
         "จากหน้าตัด",
         f"{_fmt(r.fv)} MPa"),
        ("Fv (ยอมให้)",
         "Fv = 0.40·Fy",
         f"= 0.40×{_fmt(Fy, 0)}",
         f"{_fmt(r.Fv)} MPa"),
        ("fv/Fv ≤ 1.0",
         "อัตราส่วนการใช้งาน",
         f"= {_fmt(r.fv)}/{_fmt(r.Fv)}",
         f"{_fmt(r.shear_ratio, 3)}"),
    ]
    elems.append(_formula_table(shear_rows))
    elems.append(Spacer(1, 5))

    # ── 4.5. Diagrams ─────────────────────────────────────────────────
    try:
        elems += _add_diagram_to_story(
            "4.5 แผนภาพแรงเฉือนและโมเมนต์ดัด (Diagrams)",
            _create_beam_diagram,
            result, span
        )
    except Exception:
        pass  # Silently skip if diagrams fail

    # ── 5. Deflection check ───────────────────────────────────────────
    elems.append(_SectionHeader("5.  การตรวจสอบการแอ่นตัว (Deflection Check)", W))
    elems.append(Spacer(1, 3))
    Ix_val = props.get("Ix", 0)
    defl_rows = [
        ("δ_max",
         "δ = 5wL⁴/384EI",
         f"w={_fmt(w_crit)}, L={_fmt(L_mm,0)}, EI={_fmt(E*Ix_val/1e12, 0) if Ix_val else '?'} kN·m²",
         f"{_fmt(r.delta_max)} mm"),
        ("δ_allow",
         "δ = L/360",
         f"= {_fmt(span*1000, 0)}/360",
         f"{_fmt(r.delta_allowable)} mm"),
        ("δ/δ_allow ≤ 1.0",
         "อัตราส่วนการแอ่นตัว",
         f"= {_fmt(r.delta_max)}/{_fmt(r.delta_allowable)}",
         f"{_fmt(r.deflection_ratio, 3)}"),
    ]
    elems.append(_formula_table(defl_rows, [4.5*cm, 4.5*cm, 5.5*cm, 2.2*cm]))
    elems.append(Spacer(1, 5))

    # ── 6. Summary ────────────────────────────────────────────────────
    elems.append(_SectionHeader("6.  สรุปผลการออกแบบ (Design Summary)", W))
    elems.append(Spacer(1, 5))
    sum_data = [
        ["การตรวจสอบ", "อัตราส่วน", "เกณฑ์", "ผล"],
        ["หน่วยแรงดัด (fb/Fb)", _fmt(r.stress_ratio, 3), "≤ 1.00",
         "✓ ผ่าน" if r.stress_ratio <= 1.0 else "✗ ไม่ผ่าน"],
        ["หน่วยแรงเฉือน (fv/Fv)", _fmt(r.shear_ratio, 3), "≤ 1.00",
         "✓ ผ่าน" if r.shear_ratio <= 1.0 else "✗ ไม่ผ่าน"],
        ["การแอ่นตัว (δ/δ_allow)", _fmt(r.deflection_ratio, 3), "≤ 1.00",
         "✓ ผ่าน" if r.deflection_ratio <= 1.0 else "✗ ไม่ผ่าน"],
    ]
    col_w = [7.5 * cm, 3 * cm, 3 * cm, 3.2 * cm]
    sum_tbl = Table(sum_data, colWidths=col_w)
    sum_st = _ts_base(col_w)
    for i, row in enumerate(sum_data[1:], 1):
        is_ok = "✓" in row[-1]
        bg = C_PASS_BG if is_ok else C_FAIL_BG
        fg = C_PASS    if is_ok else C_FAIL
        sum_st.add("BACKGROUND", (0, i), (-1, i), bg)
        sum_st.add("TEXTCOLOR",  (3, i), (3, i),  fg)
        sum_st.add("FONTNAME",   (3, i), (3, i),  "ThaiBd")
    sum_tbl.setStyle(sum_st)
    elems.append(sum_tbl)
    elems.append(Spacer(1, 8))
    elems.append(_PassFailBanner(r.is_ok, r.status, W))

    return elems


# ══════════════════════════════════════════════════════════════════════════════
# COLUMN REPORT
# ══════════════════════════════════════════════════════════════════════════════

def _column_elements(result, section_name: str, height: float) -> List:
    from column_design import ColumnDesignResult
    ST = _styles()
    W = 16.7 * cm
    r: ColumnDesignResult = result
    props = r.details.get("properties", {}) if r.details else {}
    A  = props.get("A",  0)
    rx = props.get("rx", 0)
    ry = props.get("ry", 0)
    Fy = getattr(r, "Fy", 0)
    E  = 2.0e5

    elems: List = []

    # 1. Section properties
    elems.append(_SectionHeader("1.  คุณสมบัติหน้าตัด (Section Properties)", W))
    elems.append(Spacer(1, 3))
    sec_props = {
        "หน้าตัด (Section)":       section_name,
        "ความสูงเสา H":             f"{_fmt(height)} m",
        "พื้นที่หน้าตัด A":         f"{_fmt(A, 0)} mm²",
        "rx":                       f"{_fmt(rx, 1)} mm",
        "ry":                       f"{_fmt(ry, 1)} mm",
        "E (เหล็ก)":                "200,000 MPa",
        "หน่วยแรงอัดที่ยอมให้ Fa": f"{_fmt(r.Fa)} MPa",
        "กรณีวิกฤต":                r.critical_load_case or "—",
    }
    elems.append(_section_props_table(sec_props))
    elems.append(Spacer(1, 5))

    # 2. Slenderness
    elems.append(_SectionHeader("2.  การตรวจสอบความชะลูด (Slenderness Check)", W))
    elems.append(Spacer(1, 3))
    Cc = math.sqrt(2 * math.pi**2 * E / Fy) if Fy > 0 else 0
    slen_rows = [
        ("KLx", "KLx = Kx·H", f"= 1.0×{_fmt(height)}", f"{_fmt(r.KLx)} m"),
        ("KLy", "KLy = Ky·H", f"= 1.0×{_fmt(height)}", f"{_fmt(r.KLy)} m"),
        ("KLx/rx", "= KLx×1000/rx", f"= {_fmt(r.KLx*1000,0)}/{_fmt(rx,1)}", f"{_fmt(r.slenderness_x, 1)}"),
        ("KLy/ry", "= KLy×1000/ry", f"= {_fmt(r.KLy*1000,0)}/{_fmt(ry,1)}", f"{_fmt(r.slenderness_y, 1)}"),
        ("(KL/r)_max", "ค่าวิกฤต ≤ 200", "max(KLx/rx, KLy/ry)", f"{_fmt(r.critical_slenderness, 1)}"),
        ("Cc", "Cc = √(2π²E/Fy)", f"= √(2π²×{_fmt(E,0)}/{_fmt(Fy,0)})", f"{_fmt(Cc, 1)}"),
    ]
    elems.append(_formula_table(slen_rows))
    elems.append(Spacer(1, 5))

    # 3. Allowable stress
    elems.append(_SectionHeader("3.  หน่วยแรงอัดที่ยอมให้ (Allowable Compressive Stress)", W))
    elems.append(Spacer(1, 3))
    KLr = r.critical_slenderness
    regime = "KL/r ≤ Cc  (Inelastic buckling)" if KLr <= Cc else "KL/r > Cc  (Elastic buckling)"
    elems.append(Paragraph(f"ช่วงการโก่ง: {regime}", ST["body"]))
    elems.append(Spacer(1, 2))
    fa_rows = [
        ("fa", "fa = P×1000/A", f"= {_fmt(r.max_axial_load)}×1000/{_fmt(A,0)}", f"{_fmt(r.fa)} MPa"),
        ("Fa", "สูตร วสท. 011038-22", "ตามค่า KL/r และ Cc", f"{_fmt(r.Fa)} MPa"),
        ("fa/Fa ≤ 1.0", "อัตราส่วนการอัด", f"= {_fmt(r.fa)}/{_fmt(r.Fa)}", f"{_fmt(r.axial_ratio, 3)}"),
    ]
    elems.append(_formula_table(fa_rows))
    elems.append(Spacer(1, 5))

    # 4. Combined loading
    elems.append(_SectionHeader("4.  การตรวจสอบแรงรวม (Combined Loading — Interaction)", W))
    elems.append(Spacer(1, 3))
    inter_rows = [
        ("P/P_allow", "P_allow = Fa×A/1000", f"= {_fmt(r.Fa)}×{_fmt(A,0)}/1000",
         f"{_fmt(r.allowable_axial_load)} kN"),
        ("Interaction", "H1: fa/Fa + fb/Fb ≤ 1.0", "แรงรวมทุกกรณี",
         f"{_fmt(r.interaction_ratio, 3)}"),
    ]
    elems.append(_formula_table(inter_rows))
    elems.append(Spacer(1, 5))

    # 5. Summary
    elems.append(_SectionHeader("5.  สรุปผลการออกแบบ (Design Summary)", W))
    elems.append(Spacer(1, 5))
    sum_data = [
        ["การตรวจสอบ", "ค่าที่เกิดขึ้น", "ค่าที่ยอมให้", "อัตราส่วน", "ผล"],
        ["ความชะลูด KL/r", _fmt(r.critical_slenderness, 1), "≤ 200", "—",
         "✓ ผ่าน" if r.critical_slenderness <= 200 else "✗ ไม่ผ่าน"],
        ["แรงอัด P (kN)", _fmt(r.max_axial_load), _fmt(r.allowable_axial_load),
         _fmt(r.axial_ratio, 3), "✓ ผ่าน" if r.axial_ratio <= 1.0 else "✗ ไม่ผ่าน"],
        ["แรงรวม (Interaction)", "—", "≤ 1.00",
         _fmt(r.interaction_ratio, 3),
         "✓ ผ่าน" if r.interaction_ratio <= 1.0 else "✗ ไม่ผ่าน"],
    ]
    col_w = [5.5*cm, 2.5*cm, 2.8*cm, 2.5*cm, 3.4*cm]
    sum_tbl = Table(sum_data, colWidths=col_w)
    st = _ts_base(col_w)
    for i, row in enumerate(sum_data[1:], 1):
        is_ok = "✓" in row[-1]
        bg = C_PASS_BG if is_ok else C_FAIL_BG
        fg = C_PASS    if is_ok else C_FAIL
        st.add("BACKGROUND", (0, i), (-1, i), bg)
        st.add("TEXTCOLOR",  (4, i), (4, i), fg)
        st.add("FONTNAME",   (4, i), (4, i), "ThaiBd")
    sum_tbl.setStyle(st)
    elems.append(sum_tbl)
    elems.append(Spacer(1, 8))
    elems.append(_PassFailBanner(r.is_ok, r.status, W))

    return elems


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def generate_beam_report(
    result,
    section_name: str,
    span: float,
    project: Optional[ProjectInfo] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Generate a PDF calculation report for a beam design result.

    Args:
        result:       BeamDesignResult from beam_design.BeamDesign.check_beam()
        section_name: e.g. "H200x200x8x12"
        span:         Beam span in metres
        project:      ProjectInfo (optional, default values used if None)
        output_path:  If given, also save PDF to this path

    Returns:
        PDF bytes
    """
    _register_fonts()
    if project is None:
        project = ProjectInfo()
    buf = io.BytesIO()
    doc = _ReportDoc(buf, project)
    story = []
    story += _cover_elements(project,
                             f"ออกแบบคานเหล็ก\n{section_name}  |  L = {_fmt(span)} m",
                             [f"คานเหล็ก {section_name}, ช่วง {_fmt(span)} m"])
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story += _beam_elements(result, section_name, span)
    doc.build(story)
    pdf_bytes = buf.getvalue()
    if output_path:
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
    return pdf_bytes


def generate_column_report(
    result,
    section_name: str,
    height: float,
    project: Optional[ProjectInfo] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """Generate PDF for a column design result."""
    _register_fonts()
    if project is None:
        project = ProjectInfo()
    buf = io.BytesIO()
    doc = _ReportDoc(buf, project)
    story = []
    story += _cover_elements(project,
                             f"ออกแบบเสาเหล็ก\n{section_name}  |  H = {_fmt(height)} m",
                             [f"เสาเหล็ก {section_name}, ความสูง {_fmt(height)} m"])
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story += _column_elements(result, section_name, height)
    doc.build(story)
    pdf_bytes = buf.getvalue()
    if output_path:
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
    return pdf_bytes


def generate_combined_report(
    modules: List[Dict],
    project: Optional[ProjectInfo] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Generate a single PDF containing multiple calculation modules.

    Args:
        modules: list of dicts, each with keys:
            - "type":    "beam" | "column"
            - "result":  design result object
            - "section": section name string
            - "param":   span (beam) or height (column) in metres
            - "title":   optional display title
        project:     ProjectInfo
        output_path: optional save path
    Returns:
        PDF bytes
    """
    _register_fonts()
    if project is None:
        project = ProjectInfo()

    from reportlab.platypus import PageBreak

    stories_list = []
    for m in modules:
        t = m.get("type", "")
        if t == "beam":
            stories_list.append(m.get("title") or
                                 f"คานเหล็ก {m['section']} L={m['param']} m")
        elif t == "column":
            stories_list.append(m.get("title") or
                                 f"เสาเหล็ก {m['section']} H={m['param']} m")

    buf = io.BytesIO()
    doc = _ReportDoc(buf, project)
    story = []
    story += _cover_elements(project, "รายการคำนวณโครงสร้างเหล็ก", stories_list)
    story.append(PageBreak())

    for i, m in enumerate(modules):
        t = m.get("type", "")
        if t == "beam":
            story += _beam_elements(m["result"], m["section"], m["param"])
        elif t == "column":
            story += _column_elements(m["result"], m["section"], m["param"])
        if i < len(modules) - 1:
            story.append(PageBreak())

    doc.build(story)
    pdf_bytes = buf.getvalue()
    if output_path:
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
    return pdf_bytes
 

# ══════════════════════════════════════════════════════════════════════════════
# DIAGRAM GENERATORS (Matplotlib → ReportLab)
# ══════════════════════════════════════════════════════════════════════════════

class _MatplotlibDiagram(Flowable):
    """Convert a matplotlib figure to a ReportLab flowable."""
    def __init__(self, fig, width=16*cm, height=10*cm, dpi=150):
        super().__init__()
        self._fig = fig
        self.width = width
        self.height = height
        self._dpi = dpi

    def wrap(self, availWidth, availHeight):
        return (self.width, self.height)

    def draw(self):
        # Save figure to BytesIO
        buf = io.BytesIO()
        self._fig.savefig(buf, format='png', dpi=self._dpi, 
                         bbox_inches='tight', pad_inches=0.1)
        buf.seek(0)
        
        # Create ReportLab Image and draw it
        from reportlab.platypus import Image
        img = Image(buf, width=self.width, height=self.height)
        img.drawOn(self.canv, self._x, self._y)


def _create_beam_diagram(result, span: float, width=16*cm, height=9*cm):
    """Create bending moment and shear force diagrams for beam."""
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 5), dpi=100)
    
    # Calculate positions
    n_points = 100
    x = np.linspace(0, span, n_points)
    
    # Get critical load case
    details = result.details if hasattr(result, 'details') else {}
    load_cases = details.get('load_cases', [])
    
    if load_cases:
        # Use the critical load case
        critical_lc = max(load_cases, key=lambda lc: lc.get('stress_ratio', 0))
        w_kN_m = critical_lc.get('w_kN_m', 0)
        V_max = w_kN_m * span / 2
        M_max = w_kN_m * span**2 / 8
        
        # Shear force diagram
        V = V_max * (1 - 2*x/span)
        ax1.fill_between(x*1000, V, 0, alpha=0.3, color='red')
        ax1.plot(x*1000, V, 'r-', linewidth=2, label=f'V_max = {V_max:.2f} kN')
        ax1.axhline(y=0, color='black', linewidth=0.5)
        ax1.set_ylabel('Shear Force (kN)', fontsize=9)
        ax1.set_title('Shear Force Diagram', fontsize=10, fontweight='bold')
        ax1.legend(fontsize=8)
        ax1.grid(True, alpha=0.3)
        
        # Bending moment diagram
        M = w_kN_m * x * (span - x) / 2
        ax2.fill_between(x*1000, M, 0, alpha=0.3, color='blue')
        ax2.plot(x*1000, M, 'b-', linewidth=2, label=f'M_max = {M_max:.2f} kN-m')
        ax2.axhline(y=0, color='black', linewidth=0.5)
        ax2.set_xlabel('Position (mm)', fontsize=9)
        ax2.set_ylabel('Bending Moment (kN-m)', fontsize=9)
        ax2.set_title('Bending Moment Diagram', fontsize=10, fontweight='bold')
        ax2.legend(fontsize=8)
        ax2.grid(True, alpha=0.3)
    else:
        # Default diagram with result values
        M_max = result.max_moment if hasattr(result, 'max_moment') else 0
        V_max = result.max_shear if hasattr(result, 'max_shear') else 0
        
        w_equiv = M_max * 8 / (span**2) if span > 0 else 0
        
        V = V_max * (1 - 2*x/span)
        M = w_equiv * x * (span - x) / 2
        
        ax1.fill_between(x*1000, V, 0, alpha=0.3, color='red')
        ax1.plot(x*1000, V, 'r-', linewidth=2)
        ax1.axhline(y=0, color='black', linewidth=0.5)
        ax1.set_ylabel('Shear Force (kN)', fontsize=9)
        ax1.set_title('Shear Force Diagram', fontsize=10, fontweight='bold')
        ax1.grid(True, alpha=0.3)
        
        ax2.fill_between(x*1000, M, 0, alpha=0.3, color='blue')
        ax2.plot(x*1000, M, 'b-', linewidth=2)
        ax2.axhline(y=0, color='black', linewidth=0.5)
        ax2.set_xlabel('Position (mm)', fontsize=9)
        ax2.set_ylabel('Bending Moment (kN-m)', fontsize=9)
        ax2.set_title('Bending Moment Diagram', fontsize=10, fontweight='bold')
        ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    return fig


def _create_purlin_diagram(result, span: float, spacing: float, slope: float, 
                           width=16*cm, height=9*cm):
    """Create bending moment and shear force diagrams for purlin."""
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    import numpy as np
    
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 5), dpi=100)
    
    # Get critical load details
    stress_check = result.get("Stress Check", {})
    details = stress_check.get("Details", {})
    w_x = details.get("w_x", 0)  # N/m
    w_y = details.get("w_y", 0)  # N/m
    
    # Convert to kN/m
    w_x_kN = w_x / 1000
    w_y_kN = w_y / 1000
    
    # Calculate positions
    n_points = 100
    x = np.linspace(0, span, n_points)
    
    # Simply supported beam formulas
    V_max = abs(w_x_kN) * span / 2
    M_max = abs(w_x_kN) * span**2 / 8
    
    # Shear force diagram
    V = V_max * (1 - 2*x/span) if span > 0 else np.zeros_like(x)
    ax1.fill_between(x*1000, V, 0, alpha=0.3, color='red')
    ax1.plot(x*1000, V, 'r-', linewidth=2, label=f'V_max = {V_max:.3f} kN')
    ax1.axhline(y=0, color='black', linewidth=0.5)
    ax1.set_ylabel('Shear Force (kN)', fontsize=9)
    ax1.set_title('Shear Force Diagram (Perpendicular to Roof)', fontsize=10, fontweight='bold')
    ax1.legend(fontsize=8)
    ax1.grid(True, alpha=0.3)
    
    # Bending moment diagram
    M = abs(w_x_kN) * x * (span - x) / 2 if span > 0 else np.zeros_like(x)
    ax2.fill_between(x*1000, M, 0, alpha=0.3, color='blue')
    ax2.plot(x*1000, M, 'b-', linewidth=2, label=f'M_max = {M_max:.3f} kN-m')
    ax2.axhline(y=0, color='black', linewidth=0.5)
    ax2.set_xlabel('Position along purlin (mm)', fontsize=9)
    ax2.set_ylabel('Bending Moment (kN-m)', fontsize=9)
    ax2.set_title('Bending Moment Diagram', fontsize=10, fontweight='bold')
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    return fig


def _add_diagram_to_story(section_title: str, diagram_func, *args, **kwargs) -> List:
    """Helper to add a diagram section to the story."""
    elems = []
    ST = _styles()
    W = 16.7 * cm
    
    elems.append(_SectionHeader(section_title, W))
    elems.append(Spacer(1, 5))
    
    try:
        fig = diagram_func(*args, **kwargs)
        if fig:
            diagram = _MatplotlibDiagram(fig, width=W, height=9*cm)
            elems.append(diagram)
            plt.close(fig)  # Clean up
    except Exception as e:
        elems.append(Paragraph(f"ไม่สามารถสร้างแผนภาพได้: {str(e)}", ST["warn"]))
    
    elems.append(Spacer(1, 5))
    return elems


# Add numpy import at the top for diagram calculations
try:
    import numpy as np
except ImportError:
    pass

# ══════════════════════════════════════════════════════════════════════════════
# PURLIN REPORT
# ══════════════════════════════════════════════════════════════════════════════

def _purlin_elements(result, section_name: str, span: float,
                     spacing: float, slope: float) -> List:
    """Build purlin calculation flowables."""
    ST = _styles()
    W = 16.7 * cm
    W_mm = span * 1000

    elems: List = []

    # ── 1. Section Properties ─────────────────────────────────────────
    elems.append(_SectionHeader("1.  คุณสมบัติหน้าตัด (Section Properties)", W))
    elems.append(Spacer(1, 3))

    inputs = result.get("Inputs", {})
    calc_stress = result.get("Calculated Stresses (MPa)", {})
    allow_stress = result.get("Allowable Stresses (MPa)", {})
    shear = result.get("Shear Check", {})
    defl_calc = result.get("Calculated Deflection (mm)", {})
    defl_allow = result.get("Allowable Deflection (mm)", {})

    sec_props = {
        "หน้าตัด (Section)": section_name,
        "ช่วงแป L": f"{_fmt(span)} m",
        "ระยะห่างแป": f"{_fmt(spacing)} m",
        "มุมลาดชัน": f"{_fmt(slope)}°",
        "Fy (กำลังคราก)": "245 MPa",
        "Sx (โมดูลัสหน้าตัด)": "ดูจากฐานข้อมูล",
        "Ix (โมเมนต์อินีเชีย)": "ดูจากฐานข้อมูล",
        "น้ำหนักตัวเอง": f"{_fmt(inputs.get('Dead Load (kPa)', 0), 3)} kPa",
    }
    elems.append(_section_props_table(sec_props))
    elems.append(Spacer(1, 5))

    # ── 2. Wind Load Data ─────────────────────────────────────────────
    elems.append(_SectionHeader("2.  น้ําหนักลม (Wind Load Calculation)", W))
    elems.append(Spacer(1, 3))

    wind_rows = [
        ("ความเร็วลมพื้นฐาน", f"{_fmt(inputs.get('Basic Wind Speed (m/s)', 0), 1)} m/s",
         "ASCE 7 / วสท.", ""),
        ("ความสูงอาคาร", f"{_fmt(inputs.get('Building Height (m)', 0), 1)} m",
         "", ""),
        ("ประเภทสภาพภูมิ", inputs.get("Exposure Category", "—"),
         "", ""),
        ("สัมประสิทธิ์ความดันลมภายใน (Cpi)",
         f"+{inputs.get('Internal Pressure Coeff (+/-)', '')}",
         "", ""),
        ("ความดันลมยกขึ้น (Uplift)",
         f"{_fmt(inputs.get('Calculated Wind Uplift (kPa)', 0), 3)} kPa",
         "qz·G·(Cp−Cpi)", ""),
        ("ความดันลมกดลง (Downward)",
         f"{_fmt(inputs.get('Calculated Wind Downward (kPa)', 0), 3)} kPa",
         "qz·G·(Cp−Cpi)", ""),
    ]
    elems.append(_formula_table(wind_rows, [5*cm, 4*cm, 5.5*cm, 2.2*cm]))
    elems.append(Spacer(1, 5))

    # ── 3. Load Combinations (Stress) ─────────────────────────────────
    elems.append(_SectionHeader("3.  การรวนน้ําหนักบรรทุก (Load Combinations — ASD)", W))
    elems.append(Spacer(1, 3))

    stress_details = result.get("Stress Check", {})
    all_cases = stress_details.get("All Cases", [])
    if all_cases:
        lc_cols = ["กรณีน้ําหนัก", "Interaction Ratio", "ผล"]
        data = [
            [Paragraph(c, ST["cell_hdr"]) for c in lc_cols]
        ]
        for lc in all_cases:
            ratio = lc.get("ratio", 0)
            bg, fg = _ratio_style(ratio)
            status_txt = "✓ ผ่าน" if ratio <= 1.0 else "✗ ไม่ผ่าน"
            row_data = [
                Paragraph(lc.get("name", ""), ST["cell"]),
                Paragraph(_fmt(ratio, 3), ST["cell_r"]),
                Paragraph(status_txt, ST["cell_c"]),
            ]
            data.append(row_data)

        col_w = [9*cm, 3.5*cm, 4.2*cm]
        tbl = Table(data, colWidths=col_w)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), C_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.4, C_BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_ROW_ALT]),
        ]))
        elems.append(tbl)
    elems.append(Spacer(1, 5))

    # ── 4. Bending Stress Check ───────────────────────────────────────
    elems.append(_SectionHeader("4.  การตรวจสอบหน่วยแรงดัด (Bending Stress Check)", W))
    elems.append(Spacer(1, 3))

    crit_lc = stress_details.get("Critical Load Case", "—")
    elems.append(_SubHeader(f"กรณีวิกฤต: {crit_lc}", W))
    elems.append(Spacer(1, 2))

    details = stress_details.get("Details", {})
    w_x = details.get("w_x", 0)
    w_y = details.get("w_y", 0)
    Mx = details.get("Mx", 0)
    My = details.get("My", 0)
    fbx = calc_stress.get("fbx", 0)
    fby = calc_stress.get("fby", 0)
    Fbx = allow_stress.get("Fbx", 0)
    Fby = allow_stress.get("Fby", 0)
    ratio = stress_details.get("Interaction Ratio", 0)

    bending_rows = [
        ("Mx (แกน X)", "Mx = wx·L²/8",
         f"= {_fmt(w_x / 1000, 3)}×{_fmt(span)}²/8",
         f"{_fmt(Mx / 1000, 3)} kN-m"),
        ("My (แกน Y)", "My = wy·L²/8",
         f"= {_fmt(w_y / 1000, 3)}×{_fmt(span)}²/8",
         f"{_fmt(My / 1000, 3)} kN-m"),
        ("fbx", "fbx = Mx/Sx",
         f"= {_fmt(Mx / 1e6, 3)}/Sx",
         f"{_fmt(fbx, 3)} MPa"),
        ("fby", "fby = My/Sy",
         f"= {_fmt(My / 1e6, 3)}/Sy",
         f"{_fmt(fby, 3)} MPa"),
        ("Fbx (ยอมให้)", "Fbx = 0.60·Fy",
         f"= 0.60×245",
         f"{_fmt(Fbx, 2)} MPa"),
        ("Fby (ยอมให้)", "Fby = 0.75·Fy",
         f"= 0.75×245",
         f"{_fmt(Fby, 2)} MPa"),
        ("Interaction", "fbx/Fbx + fby/Fby ≤ 1.0",
         f"= {_fmt(fbx/Fbx if Fbx else 0, 3)} + {_fmt(fby/Fby if Fby else 0, 3)}",
         f"{_fmt(ratio, 3)}"),
    ]
    elems.append(_formula_table(bending_rows))
    elems.append(Spacer(1, 5))

    # ── 5. Shear Check ────────────────────────────────────────────────
    elems.append(_SectionHeader("5.  การตรวจสอบหน่วยแรงเฉือน (Shear Stress Check)", W))
    elems.append(Spacer(1, 3))

    shear_rows = [
        ("V_max", "V = wx·L/2",
         f"= {_fmt(w_x / 1000, 3)}×{_fmt(span)}/2",
         f"{_fmt(shear.get('V_max (kN)', 0), 3)} kN"),
        ("fv", "fv = V/(d·tw)",
         "จากหน้าตัด",
         f"{_fmt(shear.get('fv (MPa)', 0), 3)} MPa"),
        ("Fv (ยอมให้)", "Fv = 0.40·Fy",
         f"= 0.40×245",
         f"{_fmt(shear.get('Fv (MPa)', 0), 2)} MPa"),
        ("fv/Fv ≤ 1.0", "อัตราส่วนการใช้งาน",
         f"= {_fmt(shear.get('fv (MPa)', 0), 3)}/{_fmt(shear.get('Fv (MPa)', 0), 2)}",
         f"{_fmt(shear.get('Ratio', 0), 3)}"),
    ]
    elems.append(_formula_table(shear_rows))
    elems.append(Spacer(1, 5))

    # ── 5.5. Diagrams ─────────────────────────────────────────────────
    try:
        elems += _add_diagram_to_story(
            "8.  แผนภาพแรงเฉือนและโมเมนต์ดัด (Diagrams)",
            _create_purlin_diagram,
            result, span, spacing, slope
        )
    except Exception:
        pass  # Silently skip if diagrams fail
    
    # ── 6. Deflection Check ───────────────────────────────────────────
    elems.append(_SectionHeader("6.  การตรวจสอบการแอ่นตัว (Deflection Check)", W))
    elems.append(Spacer(1, 3))

    defl_check = result.get("Deflection Check", {})
    defl_rows = [
        ("δ (Live Load)", "δ = 5wL⁴/384EI",
         "w = Live Load only",
         f"{_fmt(defl_calc.get('Live Load Vertical', 0), 3)} mm"),
        ("δ_allow (L/240)", "δ = L/240",
         f"= {_fmt(W_mm, 0)}/240",
         f"{_fmt(defl_allow.get('Live Load (L/240)', 0), 2)} mm"),
        ("δ (Total Load)", "δ = 5wL⁴/384EI",
         f"กรณี: {defl_check.get('Critical Total Load Case', '—')}",
         f"{_fmt(defl_calc.get('Total Vertical', 0), 3)} mm"),
        ("δ_allow (L/180)", "δ = L/180",
         f"= {_fmt(W_mm, 0)}/180",
         f"{_fmt(defl_allow.get('Total (L/180)', 0), 2)} mm"),
    ]
    elems.append(_formula_table(defl_rows, [4*cm, 4*cm, 6.5*cm, 2.2*cm]))
    elems.append(Spacer(1, 5))

    # ── 7. Summary ────────────────────────────────────────────────────
    elems.append(_SectionHeader("7.  สรุปผลการออกแบบ (Design Summary)", W))
    elems.append(Spacer(1, 5))

    is_ok = result.get("is_ok", False)
    sum_data = [
        ["การตรวจสอบ", "อัตราส่วน", "เกณฑ์", "ผล"],
        ["หน่วยแรงดัด (Interaction)",
         _fmt(stress_details.get("Interaction Ratio", 0), 3),
         "≤ 1.00",
         "✓ ผ่าน" if stress_details.get("Interaction Ratio", 999) <= 1.0 else "✗ ไม่ผ่าน"],
        ["หน่วยแรงเฉือน (fv/Fv)",
         _fmt(shear.get("Ratio", 0), 3), "≤ 1.00",
         "✓ ผ่าน" if shear.get("is_ok", False) else "✗ ไม่ผ่าน"],
        ["การแอ่นตัว (Live Load)",
         _fmt(defl_calc.get("Live Load Vertical", 0) /
              defl_allow.get("Live Load (L/240)", 1), 3),
         "≤ 1.00",
         "✓ ผ่าน" if "ผ่าน" in defl_check.get("Live Load", "") else "✗ ไม่ผ่าน"],
        ["การแอ่นตัว (Total Load)",
         _fmt(defl_calc.get("Total Vertical", 0) /
              defl_allow.get("Total (L/180)", 1), 3),
         "≤ 1.00",
         "✓ ผ่าน" if "ผ่าน" in defl_check.get("Total Load", "") else "✗ ไม่ผ่าน"],
    ]
    col_w = [6.5*cm, 3*cm, 3*cm, 4.2*cm]
    sum_tbl = Table(sum_data, colWidths=col_w)
    sum_st = _ts_base(col_w)
    for i, row_data in enumerate(sum_data[1:], 1):
        ok = "✓" in row_data[-1]
        bg = C_PASS_BG if ok else C_FAIL_BG
        fg = C_PASS if ok else C_FAIL
        sum_st.add("BACKGROUND", (0, i), (-1, i), bg)
        sum_st.add("TEXTCOLOR",  (3, i), (3, i), fg)
        sum_st.add("FONTNAME",   (3, i), (3, i), "ThaiBd")
    sum_tbl.setStyle(sum_st)
    elems.append(sum_tbl)
    elems.append(Spacer(1, 8))

    # ── Final Result ──────────────────────────────────────────────────
    elems.append(_PassFailBanner(is_ok, result.get("Final Result", ""), W))

    return elems


def generate_purlin_report(
    result,
    section_name: str,
    span: float,
    spacing: float,
    slope: float,
    project: Optional[ProjectInfo] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Generate a PDF calculation report for a purlin design result.

    Args:
        result:       dict from PurlinDesign.run_check()
        section_name: e.g. "C150x65x20x4.0"
        span:         Purlin span in metres
        spacing:      Purlin spacing in metres
        slope:        Roof slope in degrees
        project:      ProjectInfo (optional)
        output_path:  If given, also save PDF to this path

    Returns:
        PDF bytes
    """
    _register_fonts()
    if project is None:
        project = ProjectInfo()
    buf = io.BytesIO()
    doc = _ReportDoc(buf, project)
    story = []
    story += _cover_elements(
        project,
        f"ออกแบบแปหลังคา\n{section_name}  |  L = {_fmt(span)} m, S = {_fmt(spacing)} m",
        [f"แปหลังคา {section_name}, ช่วง {_fmt(span)} m, ระยะห่าง {_fmt(spacing)} m"],
    )
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story += _purlin_elements(result, section_name, span, spacing, slope)
    doc.build(story)
    pdf_bytes = buf.getvalue()
    if output_path:
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
    return pdf_bytes


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def export_purlin_to_excel(
    result: dict,
    section_name: str,
    span: float,
    spacing: float,
    slope: float,
    output_path: str,
    project: Optional[ProjectInfo] = None,
) -> str:
    """Export purlin calculation to Excel workbook."""
    if not EXCELPY_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()

    # Styles
    hdr_font = Font(name="Tahoma", bold=True, size=12, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    sub_font = Font(name="Tahoma", bold=True, size=11, color="1B4F72")
    body_font = Font(name="Tahoma", size=10)
    body_bold = Font(name="Tahoma", size=10, bold=True)
    pass_fill = PatternFill(start_color="D4F5E4", end_color="D4F5E4", fill_type="solid")
    fail_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    def _apply_header_style(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center_align
            cell.border = thin_border

    def _write_row(ws, row, values, bold=False):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = body_bold if bold else body_font
            cell.border = thin_border
            cell.alignment = center_align if c <= 2 else right_align

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "สรุปผลการออกแบบ"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 45

    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value = "รายการคํานวณออกแบบแปหลังคา"
    title_cell.font = Font(name="Tahoma", bold=True, size=14, color="0D1B2A")
    title_cell.alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:B2")
    sub_cell = ws1["A2"]
    sub_cell.value = f"หน้าตัด: {section_name}  |  ช่วง: {span} m  |  ระยะห่าง: {spacing} m"
    sub_cell.font = sub_font
    sub_cell.alignment = Alignment(horizontal="center")

    if project:
        ws1.merge_cells("A3:B3")
        ws1["A3"].value = f"โครงการ: {project.project_name}  |  วันที่: {project.date}"
        ws1["A3"].font = body_font
        ws1["A3"].alignment = Alignment(horizontal="center")

    row = 5
    _write_row(ws1, row, ["รายการ", "ค่า"], bold=True)
    row += 1
    inputs = result.get("Inputs", {})
    for k, v in inputs.items():
        _write_row(ws1, row, [str(k), str(v)])
        row += 1

    # Summary checks
    row += 1
    _write_row(ws1, row, ["การตรวจสอบ", "ผล"], bold=True)
    row += 1
    stress_ratio = result.get("Stress Check", {}).get("Interaction Ratio", 0)
    shear_ratio = result.get("Shear Check", {}).get("Ratio", 0)
    is_ok = result.get("is_ok", False)
    for label, val, ok in [
        ("Stress Interaction", f"{stress_ratio:.3f}", stress_ratio <= 1.0),
        ("Shear Check", f"{shear_ratio:.3f}", result.get("Shear Check", {}).get("is_ok", False)),
        ("Overall", "PASS" if is_ok else "FAIL", is_ok),
    ]:
        _write_row(ws1, row, [label, str(val)])
        for c in range(1, 3):
            ws1.cell(row=row, column=c).fill = pass_fill if ok else fail_fill
        row += 1

    # ── Sheet 2: Load Combinations ────────────────────────────────────
    ws2 = wb.create_sheet("Load Combinations")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20

    all_cases = result.get("Stress Check", {}).get("All Cases", [])
    if all_cases:
        _write_row(ws2, 1, ["Load Case", "Interaction Ratio"], bold=True)
        for i, lc in enumerate(all_cases, 2):
            ratio = lc.get("ratio", 0)
            _write_row(ws2, i, [lc.get("name", ""), f"{ratio:.3f}"])
            fill = pass_fill if ratio <= 1.0 else fail_fill
            for c in range(1, 3):
                ws2.cell(row=i, column=c).fill = fill

    # ── Sheet 3: Detailed Results ─────────────────────────────────────
    ws3 = wb.create_sheet("รายละเอียด")
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 15

    details = result.get("Stress Check", {}).get("Details", {})
    _write_row(ws3, 1, ["รายการ", "ค่า", "หน่วย"], bold=True)
    row = 2
    detail_items = [
        ("wx", details.get("w_x", 0), "N/m"),
        ("wy", details.get("w_y", 0), "N/m"),
        ("Mx", details.get("Mx", 0), "N-m"),
        ("My", details.get("My", 0), "N-m"),
        ("fbx", details.get("fbx_MPa", 0), "MPa"),
        ("fby", details.get("fby_MPa", 0), "MPa"),
    ]
    for label, val, unit in detail_items:
        _write_row(ws3, row, [label, f"{val:.4f}" if isinstance(val, float) else str(val), unit])
        row += 1

    # Save
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# TRUSS & FOOTING ELEMENTS
# ══════════════════════════════════════════════════════════════════════════════

def _truss_elements(result, section_name: str, length: float) -> List:
    ST = _styles()
    W = 16.7 * cm
    r = result
    
    elems = []
    elems.append(_SectionHeader("1.  คุณสมบัติและกำลังออกแบบ (Member Properties)", W))
    elems.append(Spacer(1, 3))
    
    sec_props = {
        "หน้าตัด (Section)": section_name,
        "ความยาว L": f"{_fmt(length)} m",
        "ประเภทแรง": r.member_type,
        "กำลังที่ยอมให้": f"{_fmt(r.allowable_force)} kN",
        "กรณีวิกฤต": r.critical_load_case,
    }
    elems.append(_section_props_table(sec_props))
    elems.append(Spacer(1, 5))
    
    elems.append(_SectionHeader("2.  สรุปผลการตรวจสอบ (Design Summary)", W))
    elems.append(Spacer(1, 5))
    sum_data = [
        ["การตรวจสอบ", "ค่าที่เกิดขึ้น", "เกณฑ์จำกัด", "อัตราส่วน", "ผล"],
        ["กำลังรับแรง", f"{_fmt(abs(r.max_force))} kN", f"{_fmt(r.allowable_force)} kN", f"{_fmt(r.ratio, 3)}", "✓ ผ่าน" if r.ratio <= 1.0 else "✗ ไม่ผ่าน"],
        ["ความชะลูด L/r", f"{_fmt(r.slenderness, 1)}", f"≤ {_fmt(r.limit_slenderness, 0)}", "—", "✓ ผ่าน" if r.slenderness <= r.limit_slenderness else "✗ ไม่ผ่าน"]
    ]
    col_w = [4.5*cm, 3*cm, 3*cm, 3*cm, 3.2*cm]
    sum_tbl = Table(sum_data, colWidths=col_w)
    st = _ts_base(col_w)
    for i in range(1, 3):
        is_ok = "✓" in sum_data[i][-1]
        st.add("BACKGROUND", (0, i), (-1, i), C_PASS_BG if is_ok else C_FAIL_BG)
    sum_tbl.setStyle(st)
    elems.append(sum_tbl)
    elems.append(Spacer(1, 8))
    elems.append(_PassFailBanner(r.is_ok, r.status, W))
    return elems

def _footing_elements(result, B: float, L: float, H: float) -> List:
    W = 16.7 * cm
    r = result
    elems = []
    elems.append(_SectionHeader("1.  การตรวจสอบแรงดันดิน (Soil Bearing Check)", W))
    elems.append(Spacer(1, 3))
    
    bearing_rows = [
        ("Area", "A = B × L", f"= {_fmt(B)} × {_fmt(L)}", f"{_fmt(r.actual_area)} m²"),
        ("q_actual", "q = P_total / A", "รวมน้ำหนักฐานราก", f"{_fmt(r.soil_pressure)} kPa"),
        ("q/qa ≤ 1.0", "อัตราส่วนแรงดันดิน", "", f"{_fmt(r.bearing_ratio, 3)}"),
    ]
    elems.append(_formula_table(bearing_rows))
    elems.append(Spacer(1, 5))
    
    elems.append(_SectionHeader("2.  การตรวจสอบแรงเฉือนคอนกรีต (Shear Check)", W))
    elems.append(Spacer(1, 3))
    shear_rows = [
        ("One-way", "Vu1 / phiVc1", "ระยะ d จากขอบเสา", f"{_fmt(r.shear_1way_ratio, 3)}"),
        ("Punching", "Vu2 / phiVc2", "รอบขอบเสา d/2", f"{_fmt(r.shear_2way_ratio, 3)}"),
    ]
    elems.append(_formula_table(shear_rows))
    elems.append(Spacer(1, 8))
    elems.append(_PassFailBanner(r.is_ok, r.status, W))
    return elems

# PUBLIC API UPDATES

def generate_truss_report(result, section_name: str, length: float, project: Optional[ProjectInfo] = None) -> bytes:
    _register_fonts()
    if project is None: project = ProjectInfo()
    buf = io.BytesIO()
    doc = _ReportDoc(buf, project)
    story = _cover_elements(project, f"ออกแบบชิ้นส่วนโครงถัก\n{section_name}", [f"ความยาวชิ้นส่วน L = {_fmt(length)} m"])
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story += _truss_elements(result, section_name, length)
    doc.build(story)
    return buf.getvalue()

def generate_footing_report(result, B: float, L: float, H: float, project: Optional[ProjectInfo] = None) -> bytes:
    _register_fonts()
    if project is None: project = ProjectInfo()
    buf = io.BytesIO()
    doc = _ReportDoc(buf, project)
    story = _cover_elements(project, f"ออกแบบฐานรากแผ่\n{_fmt(B)} x {_fmt(L)} m", [f"ความหนาฐานราก H = {_fmt(H)} m"])
    from reportlab.platypus import PageBreak
    story.append(PageBreak())
    story += _footing_elements(result, B, L, H)
    doc.build(story)
    return buf.getvalue()

